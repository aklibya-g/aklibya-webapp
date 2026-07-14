import json
import os
import csv
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import models
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from functools import wraps
from .models import Database, Capital, Expense, T1Summary, ClientBalance, FromSource, TransferType, OfficeName, InternalTransfer, DeliveryArea, SystemUser, ImportAlert
from .forms import DatabaseForm, CapitalForm, CapitalDepositForm, ExpenseForm, InternalTransferForm
from .whatsapp_parser import parse_whatsapp_text
import openpyxl

BACKUP_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "backups")


def _ensure_backup_dir():
    os.makedirs(BACKUP_DIR, exist_ok=True)


def _backup_transactions(label="auto"):
    _ensure_backup_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"transactions_{label}_{timestamp}.csv"
    filepath = os.path.join(BACKUP_DIR, filename)
    qs = Database.objects.all().order_by("id")
    if not qs.exists():
        return None
    field_names = [f.name for f in Database._meta.get_fields() if hasattr(f, 'column')]
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(field_names)
        for obj in qs:
            writer.writerow([getattr(obj, fn, '') for fn in field_names])
    return filename


def _backup_capital(label="auto"):
    _ensure_backup_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"capital_{label}_{timestamp}.csv"
    filepath = os.path.join(BACKUP_DIR, filename)
    qs = Capital.objects.all().order_by("id")
    if not qs.exists():
        return None
    field_names = [f.name for f in Capital._meta.get_fields() if hasattr(f, 'column')]
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(field_names)
        for obj in qs:
            writer.writerow([getattr(obj, fn, '') for fn in field_names])
    return filename


def _restore_transactions(filepath):
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            data = {}
            for k, v in row.items():
                if k == 'id':
                    continue
                if k in ('date', 'time'):
                    try:
                        data[k] = datetime.strptime(v, "%Y-%m-%d").date() if k == 'date' else datetime.strptime(v, "%H:%M:%S").time() if v else None
                    except (ValueError, TypeError):
                        data[k] = None
                elif k in ('transfer_amount', 'transfered_amount', 'exchange_rate'):
                    try:
                        data[k] = float(v) if v else 0
                    except ValueError:
                        data[k] = 0
                else:
                    data[k] = v if v else ''
            try:
                Database.objects.create(**data)
                count += 1
            except Exception:
                pass
        return count


def _restore_capital(filepath):
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            data = {}
            for k, v in row.items():
                if k == 'id':
                    continue
                if k in ('date', 'time'):
                    try:
                        data[k] = datetime.strptime(v, "%Y-%m-%d").date() if k == 'date' else datetime.strptime(v, "%H:%M:%S").time() if v else None
                    except (ValueError, TypeError):
                        data[k] = None
                elif k in ('cash_in', 'cash_out', 'libyan_cash', 'libyan_withdraw', 'exchange_rate'):
                    try:
                        data[k] = float(v) if v else 0
                    except ValueError:
                        data[k] = 0
                else:
                    data[k] = v if v else ''
            try:
                Capital.objects.create(**data)
                count += 1
            except Exception:
                pass
        return count


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = SystemUser.objects.filter(username=username).first()
        if user and user.check_password(password):
            if not user.is_active:
                messages.error(request, "تم تعطيل هذا الحساب. تواصل مع المدير.")
                return render(request, "login.html")
            request.session["user_id"] = user.id
            request.session["username"] = user.username
            request.session["is_admin"] = user.is_admin
            return redirect("welcome")
        messages.error(request, "اسم المستخدم أو كلمة المرور غير صحيحة")
    return render(request, "login.html")


def welcome_view(request):
    if "user_id" not in request.session:
        return redirect("login")
    return render(request, "welcome.html", {
        "username": request.session.get("username", ""),
        "is_admin": request.session.get("is_admin", False),
    })


def logout_view(request):
    request.session.flush()
    return redirect("login")


def login_required_custom(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if "user_id" not in request.session:
            return redirect("login")
        return view_func(request, *args, **kwargs)
    return wrapper


def get_current_user(request):
    user_id = request.session.get("user_id")
    if user_id:
        return SystemUser.objects.filter(id=user_id).first()
    return None


def has_perm(request, perm_name):
    user = get_current_user(request)
    if user:
        return user.has_perm(perm_name)
    return False


def _normalize_arabic(s):
    if not s:
        return ""
    replacements = {'أ': 'ا', 'إ': 'ا', 'آ': 'ا', 'ة': 'ه', 'ى': 'ي', 'ؤ': 'و', 'ئ': 'ي'}
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s.strip()

def _get_clients_with_remaining():
    clients_qs = ClientBalance.objects.all().order_by("name")
    all_transfers = list(Database.objects.values_list("sender_name", "transfered_amount"))
    transfer_map = {}
    for sn, amt in all_transfers:
        norm = _normalize_arabic(sn or "")
        transfer_map[norm] = transfer_map.get(norm, 0) + (amt or 0)

    clients = []
    total_egp = 0
    for c in clients_qs:
        deposited = Capital.objects.filter(client=c).exclude(in_type="ترحيل").aggregate(total=Sum("cash_in"))["total"] or 0
        norm_name = _normalize_arabic(c.name)
        transferred = transfer_map.get(norm_name, 0)
        remaining = deposited - transferred
        c.remaining_egp = remaining
        total_egp += remaining
        clients.append(c)
    return clients, total_egp


@login_required_custom
def home(request):
    db_count = Database.objects.count()
    cap_count = Capital.objects.count()
    exp_count = Expense.objects.count()
    client_count = ClientBalance.objects.count()

    agg = Database.objects.aggregate(
        total_lyd=Sum("transfer_amount"),
        total_egp=Sum("transfered_amount"),
    )
    db_total_lyd = agg["total_lyd"] or 0
    db_total_egp = agg["total_egp"] or 0
    avg_rate = round(db_total_egp / db_total_lyd, 3) if db_total_lyd else 0

    cap_agg = Capital.objects.aggregate(
        total_egp=Sum("cash_in"),
        total_lyd=Sum("libyan_cash"),
    )
    cap_total_egp = cap_agg["total_egp"] or 0
    cap_total_lyd = cap_agg["total_lyd"] or 0
    cap_avg_rate = round(cap_total_egp / cap_total_lyd, 3) if cap_total_lyd else 0

    exp_agg = Expense.objects.aggregate(total=Sum("amount"))
    exp_total = exp_agg["total"] or 0

    remaining_egp = cap_total_egp - db_total_egp

    clients, total_egp_clients = _get_clients_with_remaining()

    latest = list(Database.objects.order_by("-date")[:10])
    latest_capital = list(Capital.objects.select_related("client", "from_source").order_by("-date")[:10])

    return render(request, "home.html", {
        "db_count": db_count,
        "cap_count": cap_count,
        "exp_count": exp_count,
        "client_count": client_count,
        "db_total_lyd": db_total_lyd,
        "db_total_egp": db_total_egp,
        "avg_rate": avg_rate,
        "cap_total_egp": cap_total_egp,
        "cap_total_lyd": cap_total_lyd,
        "cap_avg_rate": cap_avg_rate,
        "remaining_egp": remaining_egp,
        "latest": latest,
        "latest_capital": latest_capital,
        "clients": clients,
        "total_egp_clients": total_egp_clients,
        "exp_total": exp_total,
        "today": datetime.now(),
    })


def transactions(request):
    order = request.GET.get("order", "desc")
    order_field = "-id" if order == "desc" else "id"
    qs = Database.objects.select_related("from_source", "office_name", "order_type", "transfer_type").all().order_by(order_field)
    q = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    year = request.GET.get("year", "").strip()
    month_from = request.GET.get("month_from", "").strip()
    month_to = request.GET.get("month_to", "").strip()

    if q:
        qs = qs.filter(
            models.Q(sender_name__icontains=q) |
            models.Q(receiver_name__icontains=q) |
            models.Q(order_number__icontains=q) |
            models.Q(receiver_tele__icontains=q) |
            models.Q(sender_tele__icontains=q) |
            models.Q(transfer_amount__icontains=q) |
            models.Q(exchange_rate__icontains=q) |
            models.Q(remarks__icontains=q)
        )
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if year:
        qs = qs.filter(date__year=year)
    if month_from:
        qs = qs.filter(date__month__gte=month_from)
    if month_to:
        qs = qs.filter(date__month__lte=month_to)

    agg = qs.aggregate(
        total_egp=Sum("transfered_amount"),
        total_lyd=Sum("transfer_amount"),
    )

    total_egp_val = agg["total_egp"] or 0
    total_lyd_val = agg["total_lyd"] or 0
    avg_rate = round(total_egp_val / total_lyd_val, 3) if total_lyd_val else 0
    filtered_count = qs.count()
    total_count = Database.objects.count()

    clients, total_egp_clients = _get_clients_with_remaining()

    is_filtered = bool(q or date_from or date_to or year or month_from or month_to)

    return render(request, "transactions_list.html", {
        "title": "الحوالات",
        "transactions": qs,
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
        "selected_year": year,
        "selected_month_from": month_from,
        "selected_month_to": month_to,
        "total_egp": total_egp_val,
        "total_lyd": total_lyd_val,
        "avg_rate": avg_rate,
        "clients": clients,
        "total_egp_clients": total_egp_clients,
        "order": order,
        "filtered_count": filtered_count,
        "total_count": total_count,
        "is_filtered": is_filtered,
    })


def _reverse_transaction_balance(obj):
    if "أحمد ياسين" in (obj.sender_name or ""):
        egp = obj.transfered_amount or 0
        lyd = obj.transfer_amount or 0
        if egp > 0:
            ahmed = ClientBalance.objects.filter(name__contains="أحمد ياسين").first()
            if ahmed:
                ahmed.egp_balance += egp
                ahmed.lyd_balance -= lyd
                ahmed.save()


def _reverse_capital_balance(obj):
    client = obj.client
    egp = obj.cash_in or 0
    lyd = obj.libyan_cash or 0
    if client and egp > 0:
        client.egp_balance -= egp
        client.lyd_balance -= lyd
        client.save()


def delete_transaction(request, id):
    obj = get_object_or_404(Database, id=id)
    _reverse_transaction_balance(obj)
    obj.delete()
    messages.success(request, f"✅ تم حذف العملية #{id} واستعادة الرصيد.")
    return redirect("transactions")


def delete_capital(request, id):
    obj = get_object_or_404(Capital, id=id)
    _reverse_capital_balance(obj)
    obj.delete()
    name = obj.client.name if obj.client else "-"
    messages.success(request, f"✅ تم حذف رصيد #{id} من {name} وتحديث الرصيد.")
    return redirect("capital_list")


def bulk_delete_transactions(request):
    if request.method == "POST":
        ids = request.POST.getlist("ids")
        count = 0
        for obj in Database.objects.filter(id__in=ids):
            _reverse_transaction_balance(obj)
            obj.delete()
            count += 1
        messages.success(request, f"✅ تم حذف {count} عملية واستعادة الأرصدة.")
    return redirect("transactions")


def bulk_delete_capital(request):
    if request.method == "POST":
        ids = request.POST.getlist("ids")
        count = 0
        for obj in Capital.objects.filter(id__in=ids):
            _reverse_capital_balance(obj)
            obj.delete()
            count += 1
        messages.success(request, f"✅ تم حذف {count} رصيد وتحديث الأرصدة.")
    return redirect("capital_list")


def clear_all_transactions(request):
    from django.db import transaction
    backup_file = _backup_transactions("before_clear")
    with transaction.atomic():
        ids = list(Database.objects.values_list("id", flat=True))
        for obj in Database.objects.filter(id__in=ids):
            _reverse_transaction_balance(obj)
        count = len(ids)
        Database.objects.filter(id__in=ids).delete()
    msg = f"✅ تم مسح {count} حوالة بالكامل واستعادة جميع الأرصدة."
    if backup_file:
        msg += f" (النسخة الاحتياطية: {backup_file})"
    messages.success(request, msg)
    return redirect("transactions")


def clear_all_capital(request):
    from django.db import transaction
    backup_file = _backup_capital("before_clear")
    with transaction.atomic():
        ids = list(Capital.objects.values_list("id", flat=True))
        for obj in Capital.objects.filter(id__in=ids):
            _reverse_capital_balance(obj)
        count = len(ids)
        Capital.objects.filter(id__in=ids).delete()
    msg = f"✅ تم مسح {count} رصيد بالكامل وتحديث الأرصدة."
    if backup_file:
        msg += f" (النسخة الاحتياطية: {backup_file})"
    messages.success(request, msg)
    return redirect("capital_list")


def capital_list(request):
    order = request.GET.get("order", "desc")
    order_field = "-id" if order == "desc" else "id"
    qs = Capital.objects.select_related("client", "from_source").all().order_by(order_field)
    q = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    client_id = request.GET.get("client", "").strip()
    year = request.GET.get("year", "").strip()
    month_from = request.GET.get("month_from", "").strip()
    month_to = request.GET.get("month_to", "").strip()

    if q:
        qs = qs.filter(
            models.Q(cash_in__icontains=q) |
            models.Q(libyan_cash__icontains=q) |
            models.Q(exchange_rate__icontains=q) |
            models.Q(remarks__icontains=q) |
            models.Q(client__name__icontains=q) |
            models.Q(from_source__from_field__icontains=q)
        )
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if client_id:
        qs = qs.filter(client_id=client_id)
    if year:
        qs = qs.filter(date__year=year)
    if month_from:
        qs = qs.filter(date__month__gte=month_from)
    if month_to:
        qs = qs.filter(date__month__lte=month_to)

    agg = qs.aggregate(
        total_egp=Sum("cash_in"),
        total_lyd=Sum("libyan_cash"),
    )

    total_egp_val = agg["total_egp"] or 0
    total_lyd_val = agg["total_lyd"] or 0
    avg_rate = round(total_egp_val / total_lyd_val, 3) if total_lyd_val else 0
    filtered_count = qs.count()
    total_count = Capital.objects.count()

    clients = ClientBalance.objects.all().order_by("name")

    is_filtered = bool(q or date_from or date_to or client_id or year or month_from or month_to)

    return render(request, "capital_list.html", {
        "title": "الأرصدة",
        "capitals": qs,
        "clients": clients,
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
        "selected_client": client_id,
        "selected_year": year,
        "selected_month_from": month_from,
        "selected_month_to": month_to,
        "total_egp": total_egp_val,
        "total_lyd": total_lyd_val,
        "avg_rate": avg_rate,
        "order": order,
        "filtered_count": filtered_count,
        "total_count": total_count,
        "is_filtered": is_filtered,
    })


def expenses_list(request):
    from datetime import datetime as dt
    qs = Expense.objects.all().order_by("-date")
    q = request.GET.get("q", "").strip()
    expense_type = request.GET.get("expense_type", "").strip()
    paid_by = request.GET.get("paid_by", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    date_from_month = request.GET.get("date_from_month", "").strip()
    date_to_month = request.GET.get("date_to_month", "").strip()
    amount_from = request.GET.get("amount_from", "").strip()
    amount_to = request.GET.get("amount_to", "").strip()

    if date_from_month:
        try:
            y, m = date_from_month.split("-")
            date_from = f"{y}-{m}-01"
        except:
            pass
    if date_to_month:
        try:
            y, m = date_to_month.split("-")
            import calendar
            last_day = calendar.monthrange(int(y), int(m))[1]
            date_to = f"{y}-{m}-{last_day}"
        except:
            pass

    if q:
        qs = qs.filter(
            models.Q(expense_type__icontains=q) |
            models.Q(notes__icontains=q) |
            models.Q(paid_by__icontains=q)
        )
    if expense_type:
        qs = qs.filter(expense_type__icontains=expense_type)
    if paid_by:
        qs = qs.filter(paid_by__icontains=paid_by)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if amount_from:
        qs = qs.filter(amount__gte=amount_from)
    if amount_to:
        qs = qs.filter(amount__lte=amount_to)

    expense_types = Expense.objects.values_list("expense_type", flat=True).distinct().order_by("expense_type")
    paid_by_list = Expense.objects.values_list("paid_by", flat=True).distinct().order_by("paid_by")
    total = qs.aggregate(t=Sum("amount"))["t"] or 0
    return render(request, "expenses_list.html", {
        "title": "المصروفات",
        "expenses": qs[:500],
        "q": q,
        "expense_filter": expense_type,
        "paid_by_filter": paid_by,
        "date_from": date_from,
        "date_to": date_to,
        "date_from_month": date_from_month,
        "date_to_month": date_to_month,
        "amount_from": amount_from,
        "amount_to": amount_to,
        "expense_types": expense_types,
        "paid_by_list": paid_by_list,
        "total": total,
    })


def add_expense(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ تم إضافة المصروف بنجاح.")
            return redirect("expenses_list")
    else:
        form = ExpenseForm(initial={"date": datetime.now().strftime("%Y-%m-%d")})
    expense_types = Expense.objects.values_list("expense_type", flat=True).distinct().order_by("expense_type")
    paid_by_list = Expense.objects.values_list("paid_by", flat=True).distinct().order_by("paid_by")
    return render(request, "add_expense.html", {
        "title": "إضافة مصروف",
        "form": form,
        "expense_types": expense_types,
        "paid_by_list": paid_by_list,
    })


def edit_expense(request, id):
    obj = get_object_or_404(Expense, id=id)
    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ تم تعديل المصروف بنجاح.")
            return redirect("expenses_list")
    else:
        form = ExpenseForm(instance=obj)
    expense_types = Expense.objects.values_list("expense_type", flat=True).distinct().order_by("expense_type")
    paid_by_list = Expense.objects.values_list("paid_by", flat=True).distinct().order_by("paid_by")
    return render(request, "add_expense.html", {
        "title": "تعديل مصروف",
        "form": form,
        "edit_mode": True,
        "expense_types": expense_types,
        "paid_by_list": paid_by_list,
    })


def delete_expense(request, id):
    obj = get_object_or_404(Expense, id=id)
    obj.delete()
    messages.success(request, f"✅ تم حذف المصروف #{id}.")
    return redirect("expenses_list")


def clear_all_expenses(request):
    count = Expense.objects.count()
    Expense.objects.all().delete()
    messages.success(request, f"✅ تم مسح {count} مصروف بالكامل.")
    return redirect("expenses_list")


def _import_excel(model_class, file, field_map, request):
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active
    rows_iter = iter(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in next(rows_iter, [])]
    col_index = {h: i for i, h in enumerate(header)}
    created = 0
    errors = []
    for row_idx, row in enumerate(rows_iter, start=2):
        kwargs = {}
        for excel_col, model_field in field_map.items():
            idx = col_index.get(excel_col)
            if idx is not None and idx < len(row):
                kwargs[model_field] = row[idx]
        try:
            model_class.objects.create(**kwargs)
            created += 1
        except Exception as e:
            errors.append(f"الصف {row_idx}: {e}")
    wb.close()
    if created:
        messages.success(request, f"تم استيراد {created} سجل بنجاح.")
    for err in errors:
        messages.error(request, err)


def _next_order_number(office_code=""):
    from django.db.models import Max
    from django.db.models.functions import Cast
    from django.db.models import Value, IntegerField
    today = datetime.now().strftime("%Y%m%d")
    prefix = f"خ-{today}-{office_code}-" if office_code else f"خ-{today}-"
    last = Database.objects.filter(
        order_number__startswith=prefix
    ).annotate(
        num=Cast("order_number", IntegerField())
    ).aggregate(Max("num"))["num__max"]
    if last:
        return str(last + 1)
    return f"{prefix}001"


def add_transaction(request):
    now_str = datetime.now().strftime("%Y-%m-%d")
    time_str = datetime.now().strftime("%H:%M")
    next_num = _next_order_number()
    form = DatabaseForm(initial={"order_number": next_num, "office_name": 1, "date": now_str, "time": time_str})

    clients, total_egp = _get_clients_with_remaining()
    offices = OfficeName.objects.all()
    if request.method == "POST":
        if "excel_file" in request.FILES:
            field_map = {
                "sender_name": "sender_name",
                "receiver_name": "receiver_name",
                "transfer_amount": "transfer_amount",
                "exchange_rate": "exchange_rate",
                "order_number": "order_number",
                "date": "date",
                "time": "time",
                "sender_tele": "sender_tele",
                "receiver_tele": "receiver_tele",
                "receiver_region": "receiver_region",
                "transfered_amount": "transfered_amount",
                "remarks": "remarks",
            }
            _import_excel(Database, request.FILES["excel_file"], field_map, request)
            return redirect("transactions")
        post_data = request.POST.copy()
        client_id = post_data.get("from_source", "")
        if not client_id:
            messages.error(request, "❌ يجب اختيار 'من رصيد' قبل الحفظ.")
            form = DatabaseForm(post_data)
            return render(request, "form.html", {
                "title": "إضافة حوالة جديدة",
                "form": form,
                "submit_label": "حفظ الحوالة",
                "clients": clients,
                "total_egp": total_egp,
                "offices": offices,
            })
        client_name = ClientBalance.objects.filter(id=client_id).values_list("name", flat=True).first()
        if client_name:
            from_obj, _ = FromSource.objects.get_or_create(from_field=client_name)
            post_data["from_source"] = from_obj.id
        office_id = post_data.get("office_name", "")
        if office_id:
            office = OfficeName.objects.filter(id=office_id).first()
            if office:
                post_data["order_number"] = _next_order_number(office.code)
        form = DatabaseForm(post_data)
        if form.is_valid():
            t = form.save()
            return render(request, "transaction_success.html", {"t": t})
    return render(request, "form.html", {
        "title": "إضافة حوالة جديدة",
        "form": form,
        "submit_label": "حفظ الحوالة",
        "clients": clients,
        "total_egp": total_egp,
        "offices": offices,
    })


def import_whatsapp(request):
    results = []
    raw_text = ""
    use_prev_rate = request.POST.get("use_prev_rate") or request.GET.get("use_prev_rate")
    import_date_str = request.POST.get("import_date", "")
    import_date = None
    if import_date_str:
        try:
            import_date = datetime.strptime(import_date_str, "%Y-%m-%d")
        except ValueError:
            import_date = None
    if request.method == "POST":
        use_prev_rate = request.POST.get("use_prev_rate")
        if "text_file" in request.FILES:
            raw_text = request.FILES["text_file"].read().decode("utf-8-sig")
        elif "pasted_text" in request.POST:
            raw_text = request.POST["pasted_text"]
        elif "raw_text" in request.POST:
            raw_text = request.POST["raw_text"]

        if "filtered_results" in request.POST and request.POST["filtered_results"]:
            try:
                parsed = json.loads(request.POST["filtered_results"])
            except (json.JSONDecodeError, ValueError):
                parsed = []
        elif raw_text:
            parsed = parse_whatsapp_text(raw_text)
            if use_prev_rate and parsed:
                last_rate = None
                for item in parsed:
                    if item.get("exchange_rate"):
                        last_rate = item["exchange_rate"]
                    elif last_rate:
                        item["exchange_rate"] = last_rate
                        if item.get("amount_egp") and last_rate > 0:
                            item["amount_lyd"] = round(item["amount_egp"] / last_rate, 2)
        else:
            parsed = []

        if raw_text and not parsed:
            messages.error(request, "لم يتم التعرف على أي حوالة صالحة. تأكد من تنسيق النص.")
        else:
            if "confirm" in request.POST:
                sender_client_id = request.POST.get("client_id")
                if sender_client_id:
                    ahmed = ClientBalance.objects.filter(id=sender_client_id).first()
                else:
                    ahmed = ClientBalance.objects.filter(name__contains="أحمد ياسين").first()
                from_whatsapp, _ = FromSource.objects.get_or_create(from_field="واتساب")
                created = 0
                errors = []
                created_records = []
                order_num_str = _next_order_number()
                order_num_parts = order_num_str.rsplit("-", 1)
                order_num_seq = int(order_num_parts[-1]) if len(order_num_parts) > 1 else 1
                order_num_prefix = "-".join(order_num_parts[:-1]) + "-" if len(order_num_parts) > 1 else order_num_str
                for item in parsed:
                    try:
                        amount_lyd = item["amount_lyd"]
                        amount_egp = item["amount_egp"]
                        if ahmed:
                            ahmed.egp_balance -= amount_egp
                            ahmed.lyd_balance += amount_lyd
                            ahmed.save()

                        tt_name = item.get("transfer_type", "كاش")
                        tt, _ = TransferType.objects.get_or_create(Transfer_type=tt_name)
                        rec = Database.objects.create(
                            sender_name="أحمد ياسين",
                            receiver_tele=item.get("receiver_tele") or "",
                            transfer_amount=amount_lyd,
                            exchange_rate=item["exchange_rate"],
                            transfered_amount=amount_egp,
                            transfer_type=tt,
                            order_number=f"{order_num_prefix}{order_num_seq:03d}",
                            date=import_date or datetime.now(),
                            time=datetime.now(),
                            from_source=from_whatsapp,
                        )
                        created_records.append(rec)
                        created += 1
                        order_num_seq += 1
                    except Exception as e:
                        errors.append(f"خطأ: {e}")

                if created:
                    client_label = ahmed.name if ahmed else "أحمد ياسين"
                    word = "حوالة" if created == 1 else "حوالات"
                    messages.success(request, f"✅ تم استيراد {created} {word} بنجاح وخصمها من رصيد {client_label}.")
                if not created and not errors:
                    messages.warning(request, "⚠️ لم يتم استيراد أي حوالة. تأكد من صحة البيانات.")
                if errors:
                    ImportAlert.objects.create(
                        import_type="whatsapp",
                        total_items=len(parsed),
                        success_count=created,
                        failed_count=len(errors),
                        failed_details="\n".join(errors),
                    )
                    messages.error(request, f"❌ فشل استيراد {len(errors)} من {len(parsed)} حوالة. السبب: {errors[0]}")
                    if len(errors) > 1:
                        messages.info(request, f"📋 تفاصيل الفشل ({len(errors)} رسالة):")
                        for i, err in enumerate(errors[:5], 1):
                            messages.warning(request, f"  {i}. {err}")
                        if len(errors) > 5:
                            messages.info(request, f"  ... و {len(errors) - 5} أخطاء أخرى. راجع تنبيهات الاستيراد.")
                return render(request, "import_whatsapp.html", {
                    "title": "استيراد حوالات خارجية",
                    "created_records": created_records,
                    "db_count": Database.objects.count(),
                    "db_total_egp": Database.objects.aggregate(t=Sum("transfered_amount"))["t"] or 0,
                    "use_prev_rate": use_prev_rate,
                    "whatsapp_alerts": ImportAlert.objects.filter(import_type="whatsapp")[:10],
                })
            else:
                results = parsed

    whatsapp_alerts = ImportAlert.objects.filter(import_type="whatsapp")[:10]
    return render(request, "import_whatsapp.html", {
        "title": "استيراد حوالات خارجية",
        "results": results,
        "raw_text": raw_text,
        "db_count": Database.objects.count(),
        "db_total_egp": Database.objects.aggregate(t=Sum("transfered_amount"))["t"] or 0,
        "use_prev_rate": use_prev_rate,
        "whatsapp_alerts": whatsapp_alerts,
    })


def add_capital(request):
    form = CapitalDepositForm(initial={"date": datetime.now().strftime("%Y-%m-%d")})
    if request.method == "POST":
        form = CapitalDepositForm(request.POST)
        if form.is_valid():
            client = form.cleaned_data["client"]
            amount_egp = form.cleaned_data["amount_egp"]
            exchange_rate = form.cleaned_data["exchange_rate"]
            amount_lyd = round(amount_egp / exchange_rate, 2) if exchange_rate else 0
            date = form.cleaned_data["date"]
            notes = form.cleaned_data.get("notes", "")

            client.egp_balance += amount_egp
            client.lyd_balance += amount_lyd
            client.save()

            Capital.objects.create(
                cash_in=amount_egp,
                libyan_cash=amount_lyd,
                date=date,
                in_type="إيداع",
                exchange_rate=exchange_rate,
                remarks=notes,
                client=client,
            )

            messages.success(request, f"✅ تم إيداع {amount_egp:,.2f} EGP لحساب {client.name} (مايعادل {amount_lyd:,.2f} LYD)")
            return redirect("capital_list")
    return render(request, "deposit_form.html", {
        "title": "إضافة رصيد",
        "form": form,
        "submit_label": "إيداع الرصيد",
        "back_url": "capital_list",
    })


import re
from datetime import datetime as dt


def _ar_to_western(s):
    """Convert Arabic-Indic digits (٠١٢٣٤٥٦٧٨٩) to Western digits."""
    return s.translate(str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789'))


def _is_skip_line(ln):
    """Return True if line should be skipped (phone, timestamp, noise)."""
    clean = ln.replace(" ", "").replace("-", "").replace("+", "").replace("(", "").replace(")", "")
    # Phone numbers: 10+ digits
    if re.match(r'^[\d\s\-\+\(\)]+$', clean) and len(clean) >= 10:
        return True
    # Time patterns: "3:31", "14:30", "3:31 ص", "3:31 م"
    if re.match(r'^\d{1,2}:\d{2}(\s*[صمم])?$', ln.replace(" ", "")):
        return True
    # Lines that are just dates
    if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$', ln.replace(" ", "")):
        return True
    return False


def _clean_number(raw):
    """Clean a number string: handle dot as thousands separator (500.000 = 500000)."""
    raw = raw.strip().rstrip(".")
    # If pattern is like 500.000 (3 digits after dot) -> dot is thousands separator
    if re.match(r'^\d{1,3}\.\d{3}$', raw):
        return float(raw.replace(".", ""))
    # Otherwise remove dots that are thousands separators (e.g., 1.000.000)
    parts = raw.split(".")
    if len(parts) > 2:
        return float("".join(parts))
    # Normal: replace comma with nothing, keep dot as decimal
    return float(raw.replace(",", ""))


def _extract_egp(ln):
    """Try to extract EGP amount from a single line. Returns float or None.
    Returns None if the number has 🇾🇪🇱🇾 flag (meaning it's LYD, not EGP)."""
    # If line has 🇾🇪 or 🇱🇾 flag after number AND has 'ج.م' or 'ج م', it's EGP not LYD
    # Otherwise 🇾🇪 just means "via Yemen channel", not currency - ignore it

    # Pattern 1: "استلام XXX" - extract number after استلام
    m = re.search(r'استلام\s*(\d[\d,\.]*)', ln)
    if m:
        try:
            v = _clean_number(m.group(1))
            if re.search(r'الف', ln) and v < 10000:
                v *= 1000
            elif re.search(r'مليون', ln) and v < 10000:
                v *= 1000000
            if v >= 10:
                return v
        except ValueError:
            pass

    # Pattern 2: "القيمة:XXX" or "القيمة XXX"
    m = re.search(r'القيمو?ه?ة?[:\s]*(\d[\d,\.]*)', ln)
    if m:
        try:
            v = _clean_number(m.group(1))
            if v >= 10:
                return v
        except ValueError:
            pass

    # Pattern 3: "XXX جنيه مصري" or "XXXج.م" or "XXX ج م"
    m = re.search(r'(\d[\d,\.]*)\s*ج[.\s]*م', ln)
    if m:
        try:
            v = _clean_number(m.group(1))
            if v >= 10:
                return v
        except ValueError:
            pass

    # Pattern 4: "XXXالف" or "XXX الف"
    m = re.search(r'(\d[\d,\.]*)\s*الف', ln)
    if m:
        try:
            v = _clean_number(m.group(1)) * 1000
            if v >= 10:
                return v
        except ValueError:
            pass

    # Pattern 5: "XXXمليون" or "XXX مليون"
    m = re.search(r'(\d[\d,\.]*)\s*مليون', ln)
    if m:
        try:
            v = _clean_number(m.group(1)) * 1000000
            if v >= 10:
                return v
        except ValueError:
            pass

    # Pattern 6: "حول له XXX"
    m = re.search(r'حول\s+له?\s+(\d[\d,\.]*)', ln)
    if m:
        try:
            v = _clean_number(m.group(1))
            if v >= 10:
                return v
        except ValueError:
            pass

    return None


def _extract_rate(ln):
    """Try to extract exchange rate from a single line. Returns float or None."""
    # Strip flag emojis before matching
    ln_clean = re.sub(r'[\U0001f1e6-\U0001f1ff]{2}', '', ln).strip()

    # Pattern 1: "سعر X" or "سعر: X"
    m = re.search(r'سعر[:\s]*(\d[\d,\.]*)', ln_clean)
    if m:
        r_str = m.group(1).replace(",", ".")
        try:
            v = float(r_str)
            if 3 < v < 10:
                return v
        except ValueError:
            pass

    # Pattern 2: "X♻️"
    m = re.search(r'(\d+[\.,]?\d*)\s*♻', ln_clean)
    if m:
        r_str = m.group(1).replace(",", ".")
        try:
            v = float(r_str)
            if 3 < v < 10:
                return v
        except ValueError:
            pass

    # Pattern 3: standalone rate number (when سعر is on separate line)
    m = re.match(r'^\s*(\d+[\.,]?\d*)\s*$', ln_clean)
    if m:
        r_str = m.group(1).replace(",", ".")
        try:
            v = float(r_str)
            if 3 < v < 10:
                return v
        except ValueError:
            pass

    return None


def parse_balance_lines(text):
    """Parse WhatsApp balance messages. Extracts EGP amount and rate from each block.
    
    Format: each message has 'استلام XXX' (EGP amount) and 'سعر X' (rate).
    Numbers may have attached letters like '400الف' (=400,000) or '76200🇾🇪' (=LYD).
    """
    text = _ar_to_western(text)

    # Split into blocks using WhatsApp timestamp patterns
    # Format 1: [12/07, 11:59 pm]  Format 2: [11:59 pm, 12/07/2026]
    blocks = re.split(r'\[?\d{1,2}[/-]\d{1,2},?\s*\d{1,2}:\d{2}(?:\s*(?:am|pm|ص|م))?\]?\s*|\[?\d{1,2}:\d{2}\s*(?:am|pm|ص|م)\s*,\s*\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\]?\s*', text)

    results = []
    for block in blocks:
        block = block.strip()
        if not block or len(block) < 5:
            continue

        lines = block.split('\n')
        egp = None
        rate = None
        found_istalam = False

        for li, line in enumerate(lines):
            ln = line.strip()
            if not ln:
                continue

            # Extract EGP from 'استلام' line
            if egp is None:
                egp = _extract_egp(ln)
                if egp is None and 'استلام' in ln:
                    found_istalam = True
                elif egp is not None:
                    found_istalam = False

            # If استلام found but no number on same line, check next line
            if egp is None and found_istalam and li + 1 < len(lines):
                next_ln = lines[li + 1].strip()
                if next_ln:
                    egp = _extract_egp(next_ln)
                    if egp is not None:
                        found_istalam = False

            # Extract rate from 'سعر' line or standalone number
            if rate is None:
                rate = _extract_rate(ln)

            # Check for bare number with 🇾🇪 flag - treat as EGP (not LYD)
            if egp is None:
                m = re.search(r'(\d[\d,\.]*)\s*🇾🇪', ln)
                if m:
                    try:
                        v = _clean_number(m.group(1))
                        if v >= 10:
                            egp = v
                    except ValueError:
                        pass

        if egp is None or rate is None:
            continue

        # Sanity check - skip nonsensical results
        if egp < 10:
            continue

        final_lyd = round(egp / rate, 2) if rate else 0

        results.append({
            "egp": egp,
            "rate": rate,
            "lyd": final_lyd,
        })

    return results


def import_balance(request):
    from django.shortcuts import get_object_or_404
    clients = ClientBalance.objects.all().order_by("name")
    results = []
    raw_text = ""
    selected_client_id = None
    selected_date = None
    if request.method == "POST":
        if "text_file" in request.FILES:
            raw_text = request.FILES["text_file"].read().decode("utf-8-sig")
        elif "pasted_text" in request.POST:
            raw_text = request.POST["pasted_text"]
        elif "raw_text" in request.POST:
            raw_text = request.POST["raw_text"]
        selected_client_id = request.POST.get("client_id")
        selected_date_str = request.POST.get("import_date", "").strip()
        selected_date = dt.strptime(selected_date_str, "%Y-%m-%d").date() if selected_date_str else None
        has_filtered = "filtered_results" in request.POST and request.POST.get("filtered_results")
        if has_filtered or (raw_text and selected_client_id):
            client_name = ClientBalance.objects.filter(id=selected_client_id).values_list("name", flat=True).first()
            if "filtered_results" in request.POST and request.POST["filtered_results"]:
                try:
                    parsed = json.loads(request.POST["filtered_results"])
                except (json.JSONDecodeError, ValueError):
                    parsed = []
            else:
                parsed = parse_balance_lines(raw_text)
            if not parsed:
                messages.error(request, "لم يتم التعرف على أي بيانات صالحة.")
            else:
                for item in parsed:
                    item["client_name"] = client_name or "-"
                    if not isinstance(item.get("date"), (type(None),)):
                        try:
                            if isinstance(item["date"], str):
                                item["date"] = dt.strptime(item["date"], "%Y-%m-%d").date()
                        except (ValueError, TypeError):
                            item["date"] = selected_date or dt.now().date()
                    else:
                        item["date"] = selected_date or dt.now().date()
                if "confirm" in request.POST:
                    client = get_object_or_404(ClientBalance, id=selected_client_id)
                    created = 0
                    errors = []
                    created_records = []
                    for item in parsed:
                        try:
                            client.egp_balance += item["egp"]
                            client.lyd_balance += item["lyd"]
                            rec = Capital.objects.create(
                                cash_in=item["egp"],
                                libyan_cash=item["lyd"],
                                exchange_rate=item["rate"],
                                date=item["date"],
                                in_type="إيداع",
                                client=client,
                            )
                            created_records.append(rec)
                            created += 1
                        except Exception as e:
                            errors.append(f"خطأ في سطر: EGP={item.get('egp', '?')}, السعر={item.get('rate', '?')} - {e}")
                    client.save()
                    if created:
                        messages.success(request, f"✅ تم استيراد {created} رصيد لحساب {client.name}.")
                    if not created and not errors:
                        messages.warning(request, "⚠️ لم يتم استيراد أي رصيد. تأكد من صحة البيانات.")
                    if errors:
                        ImportAlert.objects.create(
                            import_type="balance",
                            total_items=len(parsed),
                            success_count=created,
                            failed_count=len(errors),
                            failed_details="\n".join(errors),
                        )
                        messages.error(request, f"❌ فشل استيراد {len(errors)} من {len(parsed)} رصيد. السبب: {errors[0]}")
                        if len(errors) > 1:
                            messages.info(request, f"📋 تفاصيل الفشل ({len(errors)} سطر):")
                            for i, err in enumerate(errors[:5], 1):
                                messages.warning(request, f"  {i}. {err}")
                            if len(errors) > 5:
                                messages.info(request, f"  ... و {len(errors) - 5} أخطاء أخرى. راجع تنبيهات الاستيراد.")
                    return render(request, "import_balance.html", {
                        "title": "استيراد أرصدة خارجية",
                        "clients": clients,
                        "created_records": created_records,
                        "selected_client_id": selected_client_id,
                        "selected_date": selected_date,
                        "cap_count": Capital.objects.count(),
                        "cap_total_egp": Capital.objects.aggregate(t=Sum("cash_in"))["t"] or 0,
                        "balance_alerts": ImportAlert.objects.filter(import_type="balance")[:10],
                    })
                else:
                    results = parsed
    balance_alerts = ImportAlert.objects.filter(import_type="balance")[:10]
    return render(request, "import_balance.html", {
        "title": "استيراد أرصدة خارجية",
        "clients": clients,
        "results": results,
        "raw_text": raw_text,
        "selected_client_id": selected_client_id,
        "selected_date": selected_date,
        "cap_count": Capital.objects.count(),
        "cap_total_egp": Capital.objects.aggregate(t=Sum("cash_in"))["t"] or 0,
        "balance_alerts": balance_alerts,
    })


def export_transfers_excel(request):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="transfers.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الحوالات"
    headers = ["رقم الطلب", "التاريخ", "الوقت", "المرسل", "هاتف المرسل", "المستلم", "هاتف المستلم", "المبلغ مصري", "سعر الصرف", "المبلغ ليبي", "نوع التحويل", "المكتب", "منطقة المستلم", "ملاحظات"]
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    rows = Database.objects.select_related("transfer_type", "office_name").order_by("-date", "-id")
    for row_idx, t in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=t.order_number or "")
        ws.cell(row=row_idx, column=2, value=str(t.date) if t.date else "")
        ws.cell(row=row_idx, column=3, value=str(t.time) if t.time else "")
        ws.cell(row=row_idx, column=4, value=t.sender_name or "")
        ws.cell(row=row_idx, column=5, value=t.sender_tele or "")
        ws.cell(row=row_idx, column=6, value=t.receiver_name or "")
        ws.cell(row=row_idx, column=7, value=t.receiver_tele or "")
        ws.cell(row=row_idx, column=8, value=t.transfered_amount or 0)
        ws.cell(row=row_idx, column=9, value=t.exchange_rate or 0)
        ws.cell(row=row_idx, column=10, value=t.transfer_amount or 0)
        ws.cell(row=row_idx, column=11, value=t.transfer_type.Transfer_type if t.transfer_type else "")
        ws.cell(row=row_idx, column=12, value=t.office_name.office_name if t.office_name else "")
        ws.cell(row=row_idx, column=13, value=t.receiver_region or "")
        ws.cell(row=row_idx, column=14, value=t.remarks or "")
        for col in [8, 9, 10]:
            ws.cell(row=row_idx, column=col).number_format = '#,##0.000'
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 16
    ws.column_dimensions['N'].width = 18
    wb.save(response)
    return response


def export_capital_excel(request):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="capital.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الأرصدة"
    headers = ["التاريخ", "العميل", "وارد نقدي", "نقدي ليبي", "سعر الصرف", "نوع الإيداع", "ملاحظات"]
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    rows = Capital.objects.select_related("client").order_by("-date", "-id")
    for row_idx, c in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=str(c.date) if c.date else "")
        ws.cell(row=row_idx, column=2, value=c.client.name if c.client else "")
        ws.cell(row=row_idx, column=3, value=c.cash_in or 0)
        ws.cell(row=row_idx, column=4, value=c.libyan_cash or 0)
        ws.cell(row=row_idx, column=5, value=c.exchange_rate or 0)
        ws.cell(row=row_idx, column=6, value=c.in_type or "")
        ws.cell(row=row_idx, column=7, value=c.remarks or "")
        for col in [3, 4, 5]:
            ws.cell(row=row_idx, column=col).number_format = '#,##0.000'
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 22
    wb.save(response)
    return response


def import_transfers_excel(request):
    if request.method != "POST" or "excel_file" not in request.FILES:
        return redirect("import_whatsapp")
    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"])
        ws = wb.active
        created = 0
        errors = []
        ahmed = ClientBalance.objects.filter(name__contains="أحمد ياسين").first()
        from_system, _ = FromSource.objects.get_or_create(from_field="نظام")
        order_num_str = _next_order_number()
        order_num_parts = order_num_str.rsplit("-", 1)
        order_num_seq = int(order_num_parts[-1]) if len(order_num_parts) > 1 else 1
        order_num_prefix = "-".join(order_num_parts[:-1]) + "-" if len(order_num_parts) > 1 else order_num_str
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row or not row[0]:
                    continue
                amount_egp = float(row[7] or 0)
                rate = float(row[8] or 0)
                amount_lyd = float(row[9] or 0)
                if amount_egp <= 0:
                    continue
                if ahmed:
                    ahmed.egp_balance -= amount_egp
                    ahmed.lyd_balance += amount_lyd
                    ahmed.save()
                tt_name = str(row[10] or "كاش").strip()
                tt, _ = TransferType.objects.get_or_create(Transfer_type=tt_name)
                office_name_str = str(row[11] or "").strip()
                office_obj = None
                if office_name_str:
                    office_obj, _ = OfficeName.objects.get_or_create(office_name=office_name_str)
                Database.objects.create(
                    order_number=str(row[0] or f"{order_num_prefix}{order_num_seq:03d}"),
                    date=datetime.strptime(str(row[1]), "%Y-%m-%d").date() if row[1] else datetime.now().date(),
                    time=datetime.strptime(str(row[2]), "%H:%M:%S").time() if row[2] else datetime.now().time(),
                    sender_name=str(row[3] or ""),
                    sender_tele=str(row[4] or ""),
                    receiver_name=str(row[5] or ""),
                    receiver_tele=str(row[6] or ""),
                    transfered_amount=amount_egp,
                    exchange_rate=rate,
                    transfer_amount=amount_lyd,
                    transfer_type=tt,
                    office_name=office_obj,
                    receiver_region=str(row[12] or ""),
                    remarks=str(row[13] or ""),
                    from_source=from_system,
                )
                created += 1
                order_num_seq += 1
            except Exception as e:
                errors.append(f"صف {row_idx}: {e}")
        if created:
            messages.success(request, f"✅ تم استيراد {created} حوالة من ملف Excel بنجاح.")
        for err in errors:
            messages.error(request, f"❌ {err}")
    except Exception as e:
        messages.error(request, f"❌ خطأ في قراءة الملف: {e}")
    return redirect("import_whatsapp")


def import_capital_excel(request):
    if request.method != "POST" or "excel_file" not in request.FILES:
        return redirect("import_balance")
    client_id = request.POST.get("client_id")
    if not client_id:
        messages.error(request, "❌ يجب اختيار العميل أولاً.")
        return redirect("import_balance")
    client = get_object_or_404(ClientBalance, id=client_id)
    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"])
        ws = wb.active
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if not row or not row[2]:
                    continue
                egp = float(row[2] or 0)
                lyd = float(row[3] or 0)
                rate = float(row[4] or 0)
                if egp <= 0:
                    continue
                date_str = str(row[0]) if row[0] else ""
                if date_str and "-" in date_str:
                    date_val = datetime.strptime(date_str, "%Y-%m-%d").date()
                else:
                    date_val = datetime.now().date()
                client.egp_balance += egp
                client.lyd_balance += lyd
                Capital.objects.create(
                    cash_in=egp,
                    libyan_cash=lyd,
                    exchange_rate=rate,
                    date=date_val,
                    in_type=str(row[5] or "إيداع"),
                    remarks=str(row[6] or ""),
                    client=client,
                )
                created += 1
            except Exception:
                pass
        client.save()
        if created:
            messages.success(request, f"✅ تم استيراد {created} رصيد لحساب {client.name} من ملف Excel.")
    except Exception as e:
        messages.error(request, f"❌ خطأ في قراءة الملف: {e}")
    return redirect("import_balance")


@csrf_exempt
def quick_client(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if not name:
            return JsonResponse({"error": "الاسم مطلوب"}, status=400)
        client = ClientBalance.objects.create(name=name)
        return JsonResponse({"id": client.id, "name": client.name})
    return JsonResponse({"error": "method not allowed"}, status=405)


@csrf_exempt
def quick_office(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip()
        if not name:
            return JsonResponse({"error": "الاسم مطلوب"}, status=400)
        office, _ = OfficeName.objects.get_or_create(office_name=name, defaults={"code": code})
        return JsonResponse({"id": office.id, "name": office.office_name, "code": office.code})
    return JsonResponse({"error": "method not allowed"}, status=405)


def transaction_receipt(request, id):
    t = get_object_or_404(Database, id=id)
    return render(request, "transaction_receipt.html", {"t": t})


def profits_report(request):
    from datetime import datetime

    year = request.GET.get("year", "")
    month_from = request.GET.get("month_from", "")
    month_to = request.GET.get("month_to", "")

    if not year:
        year = datetime.now().year
    year = int(year)

    if month_from:
        month_from = int(month_from)
    else:
        month_from = 1

    if month_to:
        month_to = int(month_to)
    else:
        month_to = 12

    month_names = {
        1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل",
        5: "مايو", 6: "يونيو", 7: "يوليو", 8: "أغسطس",
        9: "سبتمبر", 10: "أكتوبر", 11: "نوفمبر", 12: "ديسمبر",
    }

    data = []
    total_profit = 0
    total_db_egp = 0
    total_db_lyd = 0
    total_cap_egp = 0
    total_cap_lyd = 0
    total_egp_to_lyd = 0
    total_lyd_surplus = 0

    for m in range(month_from, month_to + 1):
        tx_qs = Database.objects.filter(date__year=year, date__month=m)
        cap_qs = Capital.objects.filter(date__year=year, date__month=m)

        db_egp = tx_qs.aggregate(t=Sum("transfered_amount"))["t"] or 0
        db_lyd = tx_qs.aggregate(t=Sum("transfer_amount"))["t"] or 0
        cap_egp = cap_qs.aggregate(t=Sum("cash_in"))["t"] or 0
        cap_lyd = cap_qs.aggregate(t=Sum("libyan_cash"))["t"] or 0

        avg_rate = round(cap_egp / cap_lyd, 3) if cap_lyd else 0
        egypt_surplus_egp = cap_egp - db_egp
        egypt_surplus_lyd = egypt_surplus_egp / avg_rate if avg_rate else 0
        lyd_surplus = cap_lyd - db_lyd
        profit = egypt_surplus_lyd - lyd_surplus if avg_rate else 0

        total_db_egp += db_egp
        total_db_lyd += db_lyd
        total_cap_egp += cap_egp
        total_cap_lyd += cap_lyd
        total_egp_to_lyd += egypt_surplus_lyd
        total_lyd_surplus += lyd_surplus
        total_profit += profit

        db_avg_rate = round(db_egp / db_lyd, 3) if db_lyd else 0
        cap_avg_rate = round(cap_egp / cap_lyd, 3) if cap_lyd else 0

        data.append({
            "month": m,
            "month_name": month_names[m],
            "db_egp": db_egp,
            "db_lyd": db_lyd,
            "cap_egp": cap_egp,
            "cap_lyd": cap_lyd,
            "avg_rate": avg_rate,
            "db_avg_rate": db_avg_rate,
            "cap_avg_rate": cap_avg_rate,
            "egp_to_lyd": egypt_surplus_lyd,
            "lyd_surplus": lyd_surplus,
            "profit": profit,
        })

    months = [
        (1, "يناير"), (2, "فبراير"), (3, "مارس"), (4, "أبريل"),
        (5, "مايو"), (6, "يونيو"), (7, "يوليو"), (8, "أغسطس"),
        (9, "سبتمبر"), (10, "أكتوبر"), (11, "نوفمبر"), (12, "ديسمبر"),
    ]
    current_year = datetime.now().year
    years = list(range(current_year - 2, current_year + 3))

    total_db_avg_rate = round(total_db_egp / total_db_lyd, 3) if total_db_lyd else 0
    total_cap_avg_rate = round(total_cap_egp / total_cap_lyd, 3) if total_cap_lyd else 0

    return render(request, "profits_report.html", {
        "title": "تقرير الأرباح",
        "data": data,
        "year": year,
        "month_from": month_from,
        "month_to": month_to,
        "years": years,
        "months": months,
        "total_db_egp": total_db_egp,
        "total_db_lyd": total_db_lyd,
        "total_cap_egp": total_cap_egp,
        "total_cap_lyd": total_cap_lyd,
        "total_egp_to_lyd": total_egp_to_lyd,
        "total_lyd_surplus": total_lyd_surplus,
        "total_db_avg_rate": total_db_avg_rate,
        "total_cap_avg_rate": total_cap_avg_rate,
        "total_profit": total_profit,
    })


def export_profits_report(request):
    from datetime import datetime
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    year = int(request.GET.get("year", datetime.now().year))
    month_from = int(request.GET.get("month_from", 1))
    month_to = int(request.GET.get("month_to", 12))

    month_names = {
        1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل",
        5: "مايو", 6: "يونيو", 7: "يوليو", 8: "أغسطس",
        9: "سبتمبر", 10: "أكتوبر", 11: "نوفمبر", 12: "ديسمبر",
    }
    month_from_name = month_names.get(month_from, "")
    month_to_name = month_names.get(month_to, "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير الأرباح"

    header_font = Font(name="Cairo", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="065f46", end_color="065f46", fill_type="solid")
    ws.merge_cells("A1:L1")
    ws["A1"] = f"تقرير الأرباح — الفترة من {month_from_name} الى {month_to_name} {year}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    sub_font = Font(name="Cairo", size=10, bold=True, color="FFFFFF")
    ws.merge_cells("A2:L2")
    ws["A2"] = "شركة اليمامة المالية"
    ws["A2"].font = sub_font
    ws["A2"].fill = header_fill
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 28

    headers = ["#", "الشهر", "إيداعات (مصري)", "حوالات (مصري)", "الفرق ÷ السعر", "إيداعات (ليبي)", "حوالات (ليبي)", "الفرق الليبي", "سعر الحوالات", "سعر الارصدة", "متوسط السعر", "صافي الربح"]
    header_row = 4
    col_header_fill = PatternFill(start_color="f0fdf9", end_color="f0fdf9", fill_type="solid")
    col_header_font = Font(name="Cairo", size=10, bold=True, color="374151")
    thin_border = Border(
        left=Side(style="thin", color="e5e7eb"),
        right=Side(style="thin", color="e5e7eb"),
        top=Side(style="thin", color="e5e7eb"),
        bottom=Side(style="thin", color="e5e7eb"),
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    total_db_egp = 0
    total_db_lyd = 0
    total_cap_egp = 0
    total_cap_lyd = 0
    total_egp_to_lyd = 0
    total_lyd_surplus = 0
    total_profit = 0

    row = header_row + 1
    data_font = Font(name="Cairo", size=10)
    profit_pos_font = Font(name="Cairo", size=10, bold=True, color="059669")
    profit_neg_font = Font(name="Cairo", size=10, bold=True, color="dc2626")

    for m in range(month_from, month_to + 1):
        tx_qs = Database.objects.filter(date__year=year, date__month=m)
        cap_qs = Capital.objects.filter(date__year=year, date__month=m)

        db_egp = tx_qs.aggregate(t=Sum("transfered_amount"))["t"] or 0
        db_lyd = tx_qs.aggregate(t=Sum("transfer_amount"))["t"] or 0
        cap_egp = cap_qs.aggregate(t=Sum("cash_in"))["t"] or 0
        cap_lyd = cap_qs.aggregate(t=Sum("libyan_cash"))["t"] or 0

        avg_rate = round(cap_egp / cap_lyd, 3) if cap_lyd else 0
        egypt_surplus_egp = cap_egp - db_egp
        egypt_surplus_lyd = egypt_surplus_egp / avg_rate if avg_rate else 0
        lyd_surplus = cap_lyd - db_lyd
        profit = egypt_surplus_lyd - lyd_surplus if avg_rate else 0

        total_db_egp += db_egp
        total_db_lyd += db_lyd
        total_cap_egp += cap_egp
        total_cap_lyd += cap_lyd
        total_egp_to_lyd += egypt_surplus_lyd
        total_lyd_surplus += lyd_surplus
        total_profit += profit

        row_data = [m, month_names[m], cap_egp, db_egp, round(egypt_surplus_lyd, 2), cap_lyd, db_lyd, round(lyd_surplus, 2), round(db_egp / db_lyd, 3) if db_lyd else 0, round(cap_egp / cap_lyd, 3) if cap_lyd else 0, avg_rate, round(profit, 2)]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if col_idx == 10:
                cell.font = profit_pos_font if profit >= 0 else profit_neg_font
            if col_idx == 9:
                cell.number_format = "0.000"
        row += 1

    total_row = row
    total_fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
    total_font = Font(name="Cairo", size=10, bold=True)
    total_data = ["", "الإجمالي", total_cap_egp, total_db_egp, round(total_egp_to_lyd, 2), total_cap_lyd, total_db_lyd, round(total_lyd_surplus, 2), "—", "—", "—", round(total_profit, 2)]
    for col_idx, val in enumerate(total_data, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    col_widths = [6, 14, 18, 18, 16, 18, 18, 16, 14, 14, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"تقرير الارباح عن الفترة من {month_from_name} الى {month_to_name} سنة {year}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def edit_capital(request, id):
    obj = get_object_or_404(Capital, id=id)
    if request.method == "POST":
        form = CapitalForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ تم تعديل الرصيد بنجاح.")
            return redirect("capital_list")
    else:
        form = CapitalForm(instance=obj)
    return render(request, "deposit_form.html", {
        "title": "تعديل رصيد",
        "form": form,
        "submit_label": "تحديث الرصيد",
        "back_url": "capital_list",
    })


def edit_transaction(request, id):
    obj = get_object_or_404(Database, id=id)
    clients, total_egp = _get_clients_with_remaining()
    offices = OfficeName.objects.all()
    if request.method == "POST":
        post_data = request.POST.copy()
        client_id = post_data.get("from_source", "")
        if not client_id:
            messages.error(request, "❌ يجب اختيار 'من رصيد' قبل الحفظ.")
            form = DatabaseForm(post_data, instance=obj)
            return render(request, "form.html", {
                "title": "تعديل حوالة",
                "form": form,
                "submit_label": "تحديث الحوالة",
                "clients": clients,
                "total_egp": total_egp,
                "offices": offices,
            })
        client_name = ClientBalance.objects.filter(id=client_id).values_list("name", flat=True).first()
        if client_name:
            from_obj, _ = FromSource.objects.get_or_create(from_field=client_name)
            post_data["from_source"] = from_obj.id
        form = DatabaseForm(post_data, instance=obj)
        if form.is_valid():
            t = form.save()
            return render(request, "transaction_success.html", {"t": t})
    else:
        form = DatabaseForm(instance=obj)
    return render(request, "form.html", {
        "title": "تعديل حوالة",
        "form": form,
        "submit_label": "تحديث الحوالة",
        "clients": clients,
        "total_egp": total_egp,
        "offices": offices,
    })


def dismiss_carryover(request, period):
    if request.method == "POST":
        request.session["carryover_reminder_dismissed"] = period
    return JsonResponse({"ok": True})


def carryover_balance(request):
    clients, total_egp = _get_clients_with_remaining()
    source_year = request.GET.get("source_year") or request.POST.get("source_year") or ""
    source_month = request.GET.get("source_month") or request.POST.get("source_month") or ""
    target_month = request.GET.get("target_month") or request.POST.get("target_month") or ""
    person_id = request.GET.get("person_id") or request.POST.get("person_id") or ""

    selected_client = None
    source_egp = 0
    source_lyd = 0
    source_rate = 0
    carryover_exists = False

    if person_id:
        selected_client = ClientBalance.objects.filter(id=person_id).first()

    if request.method == "POST" and person_id and source_month and source_year and target_month:
        if selected_client:
            sy = int(source_year)
            sm = int(source_month)
            tm = int(target_month)
            carryover_exists = Capital.objects.filter(
                client=selected_client, in_type="ترحيل",
                date__year=sy, date__month=tm,
            ).exists()
            if not carryover_exists:
                dep = Capital.objects.filter(client=selected_client).exclude(in_type="ترحيل").aggregate(total=Sum("cash_in"))["total"] or 0
                used = Database.objects.filter(from_source__from_field__iexact=selected_client.name).aggregate(total=Sum("transfered_amount"))["total"] or 0
                egp = dep - used
                lyd_cap = Capital.objects.filter(client=selected_client).exclude(in_type="ترحيل").aggregate(total=Sum("libyan_cash"))["total"] or 0
                avg_rate = round(dep / lyd_cap, 3) if lyd_cap else 1
                lyd = round(egp / avg_rate, 2) if avg_rate else 0
                Capital.objects.create(
                    cash_in=-egp,
                    libyan_cash=-lyd,
                    date=datetime(sy, sm, 1),
                    in_type="ترحيل",
                    exchange_rate=avg_rate,
                    remarks=f"ترحيل سالب من شهر {sm} (خصم)",
                    client=selected_client,
                )
                Capital.objects.create(
                    cash_in=egp,
                    libyan_cash=lyd,
                    date=datetime(sy, tm, 1),
                    in_type="ترحيل",
                    exchange_rate=avg_rate,
                    remarks=f"ترحيل موجب إلى شهر {tm} (إضافة)",
                    client=selected_client,
                )
                messages.success(request, f"✅ تم ترحيل {egp:,.2f} EGP لحساب {selected_client.name} من شهر {sm} إلى شهر {tm}.")
                return redirect("carryover_balance")
    elif selected_client and source_year and source_month:
        sy = int(source_year)
        sm = int(source_month)
        tm = int(target_month) if target_month else (sm + 1 if sm < 12 else 1)
        carryover_exists = Capital.objects.filter(
            client=selected_client, in_type="ترحيل",
            date__year=sy, date__month=tm,
        ).exists()
        dep = Capital.objects.filter(client=selected_client).exclude(in_type="ترحيل").aggregate(total=Sum("cash_in"))["total"] or 0
        used = Database.objects.filter(from_source__from_field__iexact=selected_client.name).aggregate(total=Sum("transfered_amount"))["total"] or 0
        source_egp = dep - used
        lyd_cap = Capital.objects.filter(client=selected_client).exclude(in_type="ترحيل").aggregate(total=Sum("libyan_cash"))["total"] or 0
        avg_rate = round(dep / lyd_cap, 3) if lyd_cap else 1
        source_rate = avg_rate
        source_lyd = round(source_egp / source_rate, 2) if source_rate else 0

    return render(request, "carryover.html", {
        "title": "ترحيل الرصيد",
        "clients": clients,
        "selected_client": selected_client,
        "source_egp": source_egp,
        "source_lyd": source_lyd,
        "source_rate": source_rate,
        "carryover_exists": carryover_exists,
    })


def _next_internal_order():
    from datetime import date
    today = date.today()
    prefix = f"د-{today.strftime('%Y%m%d')}"
    last = InternalTransfer.objects.filter(order_number__startswith=prefix).order_by("-id").first()
    if last:
        seq = int(last.order_number[-3:]) + 1
    else:
        seq = 1
    return f"{prefix}-{seq:03d}"


@csrf_exempt
def add_internal_transfer(request):
    if request.method == "POST":
        form = InternalTransferForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.order_number = _next_internal_order()
            if not obj.date:
                obj.date = datetime.now().date()
            if not obj.time:
                obj.time = datetime.now().time()
            obj.save()
            messages.success(request, f"تم إرسال الحوالة داخلية #{obj.order_number} بنجاح")
            return redirect("internal_transfer_success", id=obj.id)
    else:
        form = InternalTransferForm(initial={
            "date": datetime.now().date(),
            "time": datetime.now().strftime("%H:%M"),
        })
    offices = OfficeName.objects.all()
    return render(request, "internal_transfer_form.html", {
        "form": form,
        "offices": offices,
        "title": "إرسال حوالة داخلية",
    })


def internal_transfer_success(request, id):
    obj = get_object_or_404(InternalTransfer, id=id)
    return render(request, "internal_transfer_success.html", {"t": obj})


def internal_transfer_list(request):
    user = get_current_user(request)
    q = request.GET.get("q", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    status = request.GET.get("status", "").strip()

    qs = InternalTransfer.objects.all()
    if user and not user.is_admin and user.office_name:
        qs = qs.filter(
            models.Q(sender_office__office_name=user.office_name) |
            models.Q(receiver_office__office_name=user.office_name)
        )
    if q:
        qs = qs.filter(
            models.Q(order_number__icontains=q) |
            models.Q(sender_name__icontains=q) |
            models.Q(receiver_name__icontains=q) |
            models.Q(sender_tele__icontains=q) |
            models.Q(receiver_tele__icontains=q)
        )
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if status:
        qs = qs.filter(status=status)

    agg = qs.aggregate(
        total_sent=Sum("sent_amount"),
        total_commission=Sum("office_commission"),
        total_received=Sum("received_amount"),
        count=Count("id"),
    )

    return render(request, "internal_transfer_list.html", {
        "transfers": qs[:500],
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
        "selected_status": status,
        "total_sent": agg["total_sent"] or 0,
        "total_commission": agg["total_commission"] or 0,
        "total_received": agg["total_received"] or 0,
        "count": agg["count"] or 0,
        "title": "الحوالات الداخلية",
    })


def internal_transfer_receipt(request, id):
    obj = get_object_or_404(InternalTransfer, id=id)
    return render(request, "internal_transfer_receipt.html", {"t": obj})


@csrf_exempt
def bulk_delete_internal(request):
    if request.method == "POST":
        ids = request.POST.getlist("ids")
        if ids:
            InternalTransfer.objects.filter(id__in=ids).delete()
            messages.success(request, f"تم حذف {len(ids)} حوالة")
        else:
            InternalTransfer.objects.all().delete()
            messages.success(request, "تم حذف جميع الحوالات الداخلية")
    return redirect("internal_transfer_list")


@csrf_exempt
def quick_area(request):
    if request.method == "POST":
        import json
        data = json.loads(request.body)
        name = data.get("name", "").strip()
        if name:
            obj, created = DeliveryArea.objects.get_or_create(name=name)
            return JsonResponse({"id": obj.id, "name": obj.name, "created": created})
    elif request.method == "GET":
        areas = list(DeliveryArea.objects.values("id", "name"))
        return JsonResponse({"areas": areas})
    return JsonResponse({"error": "invalid"}, status=400)


@csrf_exempt
def update_internal_status(request, id):
    if request.method == "POST":
        import json
        data = json.loads(request.body)
        status = data.get("status", "")
        valid = ["pending", "sent", "received", "cancelled"]
        if status in valid:
            InternalTransfer.objects.filter(id=id).update(status=status)
            return JsonResponse({"ok": True, "status": status})
    return JsonResponse({"error": "invalid"}, status=400)


@login_required_custom
def users_list(request):
    user = get_current_user(request)
    if not user or not user.has_perm("perm_users"):
        messages.error(request, "ليس لديك صلاحية الوصول لهذه الصفحة")
        return redirect("home")
    users = SystemUser.objects.all().order_by("-is_admin", "username")
    return render(request, "users_list.html", {"users": users})


@login_required_custom
def add_user(request):
    user = get_current_user(request)
    if not user or not user.has_perm("perm_users"):
        messages.error(request, "ليس لديك صلاحية الوصول لهذه الصفحة")
        return redirect("home")
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        office_name = request.POST.get("office_name", "").strip()
        is_admin = request.POST.get("is_admin") == "1"
        if not username or not password:
            messages.error(request, "اسم المستخدم وكلمة المرور مطلوبان")
        elif SystemUser.objects.filter(username=username).exists():
            messages.error(request, "اسم المستخدم موجود بالفعل")
        else:
            new_user = SystemUser(
                username=username,
                office_name=office_name,
                is_admin=is_admin,
                perm_home=request.POST.get("perm_home") == "1",
                perm_transactions=request.POST.get("perm_transactions") == "1",
                perm_add_transaction=request.POST.get("perm_add_transaction") == "1",
                perm_edit_transaction=request.POST.get("perm_edit_transaction") == "1",
                perm_delete_transaction=request.POST.get("perm_delete_transaction") == "1",
                perm_capital=request.POST.get("perm_capital") == "1",
                perm_add_capital=request.POST.get("perm_add_capital") == "1",
                perm_edit_capital=request.POST.get("perm_edit_capital") == "1",
                perm_delete_capital=request.POST.get("perm_delete_capital") == "1",
                perm_expenses=request.POST.get("perm_expenses") == "1",
                perm_add_expense=request.POST.get("perm_add_expense") == "1",
                perm_edit_expense=request.POST.get("perm_edit_expense") == "1",
                perm_delete_expense=request.POST.get("perm_delete_expense") == "1",
                perm_import_whatsapp=request.POST.get("perm_import_whatsapp") == "1",
                perm_import_balance=request.POST.get("perm_import_balance") == "1",
                perm_internal_transfer=request.POST.get("perm_internal_transfer") == "1",
                perm_profits_report=request.POST.get("perm_profits_report") == "1",
                perm_users=request.POST.get("perm_users") == "1",
            )
            new_user.set_password(password)
            new_user.save()
            messages.success(request, f"✅ تم إضافة المستخدم {username} بنجاح")
            return redirect("users_list")
    return render(request, "user_form.html", {"mode": "add"})


@login_required_custom
def edit_user(request, id):
    current_user = get_current_user(request)
    if not current_user or not current_user.has_perm("perm_users"):
        messages.error(request, "ليس لديك صلاحية الوصول لهذه الصفحة")
        return redirect("home")
    user_obj = get_object_or_404(SystemUser, id=id)
    if request.method == "POST":
        user_obj.office_name = request.POST.get("office_name", "").strip()
        user_obj.is_admin = request.POST.get("is_admin") == "1"
        user_obj.perm_home = request.POST.get("perm_home") == "1"
        user_obj.perm_transactions = request.POST.get("perm_transactions") == "1"
        user_obj.perm_add_transaction = request.POST.get("perm_add_transaction") == "1"
        user_obj.perm_edit_transaction = request.POST.get("perm_edit_transaction") == "1"
        user_obj.perm_delete_transaction = request.POST.get("perm_delete_transaction") == "1"
        user_obj.perm_capital = request.POST.get("perm_capital") == "1"
        user_obj.perm_add_capital = request.POST.get("perm_add_capital") == "1"
        user_obj.perm_edit_capital = request.POST.get("perm_edit_capital") == "1"
        user_obj.perm_delete_capital = request.POST.get("perm_delete_capital") == "1"
        user_obj.perm_expenses = request.POST.get("perm_expenses") == "1"
        user_obj.perm_add_expense = request.POST.get("perm_add_expense") == "1"
        user_obj.perm_edit_expense = request.POST.get("perm_edit_expense") == "1"
        user_obj.perm_delete_expense = request.POST.get("perm_delete_expense") == "1"
        user_obj.perm_import_whatsapp = request.POST.get("perm_import_whatsapp") == "1"
        user_obj.perm_import_balance = request.POST.get("perm_import_balance") == "1"
        user_obj.perm_internal_transfer = request.POST.get("perm_internal_transfer") == "1"
        user_obj.perm_profits_report = request.POST.get("perm_profits_report") == "1"
        user_obj.perm_users = request.POST.get("perm_users") == "1"
        new_pass = request.POST.get("password", "").strip()
        if new_pass:
            user_obj.set_password(new_pass)
        user_obj.save()
        messages.success(request, f"✅ تم تعديل بيانات المستخدم {user_obj.username} بنجاح")
        return redirect("users_list")
    return render(request, "user_form.html", {"mode": "edit", "user_obj": user_obj})


@login_required_custom
def delete_user(request, id):
    current_user = get_current_user(request)
    if not current_user or not current_user.has_perm("perm_users"):
        messages.error(request, "ليس لديك صلاحية الوصول لهذه الصفحة")
        return redirect("home")
    user_obj = get_object_or_404(SystemUser, id=id)
    if user_obj.is_admin:
        messages.error(request, "لا يمكن حذف مدير النظام")
        return redirect("users_list")
    user_obj.delete()
    messages.success(request, f"تم حذف المستخدم {user_obj.username}")
    return redirect("users_list")


@login_required_custom
def messages_inbox(request):
    user = get_current_user(request)
    if not user:
        return redirect("login")
    from .models import Message
    msgs = Message.objects.filter(receiver=user).order_by("-created_at")
    unread_count = msgs.filter(is_read=False).count()
    return render(request, "messages_inbox.html", {
        "messages_list": msgs,
        "unread_count": unread_count,
    })


@login_required_custom
def messages_sent(request):
    user = get_current_user(request)
    if not user:
        return redirect("login")
    from .models import Message
    msgs = Message.objects.filter(sender=user).order_by("-created_at")
    return render(request, "messages_sent.html", {"messages_list": msgs})


@login_required_custom
def message_compose(request):
    user = get_current_user(request)
    if not user:
        return redirect("login")
    from .models import Message, SystemUser
    users = SystemUser.objects.exclude(id=user.id).order_by("username")
    if request.method == "POST":
        receiver_id = request.POST.get("receiver")
        subject = request.POST.get("subject", "").strip()
        body = request.POST.get("body", "").strip()
        if receiver_id and subject and body:
            receiver = SystemUser.objects.filter(id=receiver_id).first()
            if receiver:
                Message.objects.create(sender=user, receiver=receiver, subject=subject, body=body)
                messages.success(request, f"تم إرسال الرسالة إلى {receiver.username}")
                return redirect("messages_sent")
            else:
                messages.error(request, "المستلم غير موجود")
        else:
            messages.error(request, "يرجى ملء جميع الحقول")
    return render(request, "message_compose.html", {"users": users})


@login_required_custom
def message_detail(request, id):
    user = get_current_user(request)
    if not user:
        return redirect("login")
    from .models import Message
    msg = get_object_or_404(Message, id=id)
    if msg.receiver != user and msg.sender != user:
        messages.error(request, "ليس لديك صلاحية لعرض هذه الرسالة")
        return redirect("messages_inbox")
    if msg.receiver == user and not msg.is_read:
        msg.is_read = True
        msg.save()
    if request.method == "POST" and msg.receiver == user:
        reply_text = request.POST.get("reply", "").strip()
        if reply_text:
            from django.utils import timezone
            msg.reply = reply_text
            msg.is_replied = True
            msg.reply_date = timezone.now()
            msg.save()
            messages.success(request, "تم إرسال الرد بنجاح")
            return redirect("messages_inbox")
    return render(request, "message_detail.html", {"msg": msg})


@login_required_custom
def message_delete(request, id):
    user = get_current_user(request)
    if not user:
        return redirect("login")
    from .models import Message
    msg = get_object_or_404(Message, id=id)
    if msg.sender == user or msg.receiver == user:
        msg.delete()
        messages.success(request, "تم حذف الرسالة")
    return redirect("messages_inbox")


@login_required_custom
def api_unread_count(request):
    user = get_current_user(request)
    if not user:
        return JsonResponse({"count": 0, "new_messages": []})
    from .models import Message
    unread = Message.objects.filter(receiver=user, is_read=False).order_by("-created_at")
    count = unread.count()
    new_msgs = []
    last_seen_id = request.session.get("last_seen_msg_id", 0)
    for m in unread:
        if m.id > last_seen_id:
            new_msgs.append({
                "id": m.id,
                "sender": m.sender.username,
                "subject": m.subject,
                "body": m.body[:100],
                "time": m.created_at.strftime("%H:%M"),
            })
    if unread.exists():
        request.session["last_seen_msg_id"] = unread.first().id
    return JsonResponse({"count": count, "new_messages": new_msgs})


@login_required_custom
def api_new_internal_transfers(request):
    user = get_current_user(request)
    if not user:
        return JsonResponse({"count": 0, "new_transfers": []})
    qs = InternalTransfer.objects.all()
    if not user.is_admin and user.office_name:
        qs = qs.filter(
            models.Q(sender_office__office_name=user.office_name) |
            models.Q(receiver_office__office_name=user.office_name)
        )
    new_transfers = []
    last_seen_id = request.session.get("last_seen_transfer_id", 0)
    recent = qs.filter(id__gt=last_seen_id).order_by("-id")[:10]
    for t in recent:
        if t.id > last_seen_id:
            office_name = ""
            if t.receiver_office and t.receiver_office.office_name == user.office_name:
                office_name = t.sender_office.office_name if t.sender_office else ""
                direction = "من"
            else:
                office_name = t.receiver_office.office_name if t.receiver_office else ""
                direction = "إلى"
            new_transfers.append({
                "id": t.id,
                "order_number": t.order_number,
                "sender_name": t.sender_name,
                "receiver_name": t.receiver_name,
                "sent_amount": t.sent_amount,
                "office_name": office_name,
                "direction": direction,
                "date": t.date.strftime("%Y-%m-%d"),
            })
    all_ids = list(qs.values_list("id", flat=True)[:100])
    if all_ids:
        request.session["last_seen_transfer_id"] = max(all_ids)
    return JsonResponse({"count": len(new_transfers), "new_transfers": new_transfers})


@login_required_custom
def export_internal_excel(request):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="internal_transfers.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الحوالات الداخلية"
    headers = ["رقم الحوالة", "التاريخ", "الوقت", "المرسل", "هاتف المرسل", "مكتب الإرسال", "المستلم", "هاتف المستلم", "مكتب الاستقبال", "المنطقة", "المبلغ المرسل", "العمولة", "المبلغ المستلم", "الحالة", "ملاحظات المرسل", "ملاحظات المستلم"]
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="0d9488", end_color="0d9488", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    rows = InternalTransfer.objects.all().order_by("-id")
    status_map = {"pending": "قيد الإرسال", "sent": "تم الإرسال", "received": "تم الاستلام", "cancelled": "ملغي"}
    for row_idx, t in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=t.order_number or "")
        ws.cell(row=row_idx, column=2, value=str(t.date) if t.date else "")
        ws.cell(row=row_idx, column=3, value=str(t.time) if t.time else "")
        ws.cell(row=row_idx, column=4, value=t.sender_name or "")
        ws.cell(row=row_idx, column=5, value=t.sender_tele or "")
        ws.cell(row=row_idx, column=6, value=t.sender_office.office_name if t.sender_office else "")
        ws.cell(row=row_idx, column=7, value=t.receiver_name or "")
        ws.cell(row=row_idx, column=8, value=t.receiver_tele or "")
        ws.cell(row=row_idx, column=9, value=t.receiver_office.office_name if t.receiver_office else "")
        ws.cell(row=row_idx, column=10, value=t.delivery_area_name or (t.delivery_area.name if t.delivery_area else ""))
        ws.cell(row=row_idx, column=11, value=t.sent_amount or 0)
        ws.cell(row=row_idx, column=12, value=t.office_commission or 0)
        ws.cell(row=row_idx, column=13, value=t.received_amount or 0)
        ws.cell(row=row_idx, column=14, value=status_map.get(t.status, t.status))
        ws.cell(row=row_idx, column=15, value=t.sender_notes or "")
        ws.cell(row=row_idx, column=16, value=t.receiver_notes or "")
        for col in [11, 12, 13]:
            ws.cell(row=row_idx, column=col).number_format = '#,##0.000'
    for i, w in enumerate([14, 12, 10, 16, 14, 16, 16, 14, 16, 16, 14, 12, 14, 12, 18, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(response)
    return response


@login_required_custom
def import_internal_excel(request):
    if request.method != "POST" or "excel_file" not in request.FILES:
        return redirect("internal_transfer_list")
    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"])
        ws = wb.active
        created = 0
        errors = []
        status_map = {"قيد الإرسال": "pending", "تم الإرسال": "sent", "تم الاستلام": "received", "ملغي": "cancelled", "pending": "pending", "sent": "sent", "received": "received", "cancelled": "cancelled"}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row or not row[0]:
                    continue
                order_number = str(row[0] or "").strip()
                if not order_number:
                    errors.append(f"صف {row_idx}: رقم الحوالة فارغ")
                    continue
                if InternalTransfer.objects.filter(order_number=order_number).exists():
                    continue
                sender_office = None
                if row[5]:
                    sender_office, _ = OfficeName.objects.get_or_create(office_name=str(row[5]).strip())
                receiver_office = None
                if row[8]:
                    receiver_office, _ = OfficeName.objects.get_or_create(office_name=str(row[8]).strip())
                delivery_area = None
                delivery_area_name = str(row[9] or "").strip()
                if delivery_area_name:
                    delivery_area, _ = DeliveryArea.objects.get_or_create(name=delivery_area_name)
                status_str = str(row[13] or "قيد الإرسال").strip()
                status = status_map.get(status_str, "pending")
                sent = float(row[10] or 0)
                comm = float(row[11] or 0)
                received = float(row[12] or 0)
                if not received and sent and comm:
                    received = sent - comm
                InternalTransfer.objects.create(
                    order_number=order_number,
                    date=datetime.strptime(str(row[1]), "%Y-%m-%d").date() if row[1] else datetime.now().date(),
                    time=datetime.strptime(str(row[2]), "%H:%M:%S").time() if row[2] else datetime.now().time(),
                    sender_name=str(row[3] or ""),
                    sender_tele=str(row[4] or ""),
                    sender_office=sender_office,
                    receiver_name=str(row[6] or ""),
                    receiver_tele=str(row[7] or ""),
                    receiver_office=receiver_office,
                    delivery_area=delivery_area,
                    delivery_area_name=delivery_area_name if not delivery_area else "",
                    sent_amount=sent,
                    office_commission=comm,
                    received_amount=received,
                    status=status,
                    sender_notes=str(row[14] or ""),
                    receiver_notes=str(row[15] or ""),
                )
                created += 1
            except Exception as e:
                errors.append(f"صف {row_idx}: {e}")
        if created:
            messages.success(request, f"تم استيراد {created} حوالة داخلية من ملف Excel بنجاح.")
        for err in errors:
            messages.error(request, err)
    except Exception as e:
        messages.error(request, f"خطأ في قراءة الملف: {e}")
    return redirect("internal_transfer_list")


@login_required_custom
def export_expenses_excel(request):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="expenses.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "المصروفات"
    headers = ["التاريخ", "نوع المصروف", "المبلغ", "المدفوع بواسطة", "ملاحظات"]
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="d97706", end_color="d97706", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    rows = Expense.objects.all().order_by("-date", "-id")
    for row_idx, e in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=str(e.date) if e.date else "")
        ws.cell(row=row_idx, column=2, value=e.expense_type or "")
        ws.cell(row=row_idx, column=3, value=e.amount or 0)
        ws.cell(row=row_idx, column=4, value=e.paid_by or "")
        ws.cell(row=row_idx, column=5, value=e.notes or "")
        ws.cell(row=row_idx, column=3).number_format = '#,##0.000'
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 24
    wb.save(response)
    return response


@login_required_custom
def import_expenses_excel(request):
    if request.method != "POST" or "excel_file" not in request.FILES:
        return redirect("expenses_list")
    try:
        wb = openpyxl.load_workbook(request.FILES["excel_file"])
        ws = wb.active
        created = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row or not row[0]:
                    continue
                amount = float(row[2] or 0)
                if amount <= 0:
                    continue
                Expense.objects.create(
                    date=datetime.strptime(str(row[0]), "%Y-%m-%d").date() if row[0] else datetime.now().date(),
                    expense_type=str(row[1] or "أخرى").strip(),
                    amount=amount,
                    paid_by=str(row[3] or "").strip(),
                    notes=str(row[4] or "").strip(),
                )
                created += 1
            except Exception as e:
                errors.append(f"صف {row_idx}: {e}")
        if created:
            messages.success(request, f"تم استيراد {created} مصروف من ملف Excel بنجاح.")
        for err in errors:
            messages.error(request, err)
    except Exception as e:
        messages.error(request, f"خطأ في قراءة الملف: {e}")
    return redirect("expenses_list")


@login_required_custom
def toggle_maintenance(request):
    user = get_current_user(request)
    if not user or not user.is_admin:
        messages.error(request, "ليس لديك صلاحية")
        return redirect("home")
    if request.method == "POST":
        from .models import SystemSetting
        setting = SystemSetting.get()
        setting.maintenance_mode = not setting.maintenance_mode
        setting.save()
        status = "مفعل" if setting.maintenance_mode else "معطل"
        messages.success(request, f"تم {status} وضع الصيانة")
    return redirect("home")


@login_required_custom
def toggle_user_active(request, id):
    user = get_current_user(request)
    if not user or not user.is_admin:
        messages.error(request, "ليس لديك صلاحية")
        return redirect("home")
    target = get_object_or_404(SystemUser, id=id)
    if target.id == user.id:
        messages.error(request, "لا يمكن تعطيل حسابك الخاص")
        return redirect("users_list")
    if request.method == "POST":
        target.is_active = not target.is_active
        target.save()
        status = "تفعيل" if target.is_active else "تعطيل"
        messages.success(request, f"تم {status} حساب {target.username}")
    return redirect("users_list")


@login_required_custom
def import_alerts(request):
    user = get_current_user(request)
    alerts = ImportAlert.objects.all()
    if not user or not user.is_admin:
        alerts = alerts[:0]
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "mark_read":
            alert_id = request.POST.get("alert_id")
            if alert_id:
                ImportAlert.objects.filter(id=alert_id).update(is_read=True)
            return redirect("import_alerts")
        elif action == "mark_all_read":
            ImportAlert.objects.filter(is_read=False).update(is_read=True)
            messages.success(request, "تم تعيين الكل كمقروء")
            return redirect("import_alerts")
        elif action == "clear_all":
            ImportAlert.objects.all().delete()
            messages.success(request, "تم حذف جميع التنبيهات")
            return redirect("import_alerts")
    unread_count = ImportAlert.objects.filter(is_read=False).count()
    return render(request, "import_alerts.html", {
        "title": "تنبيهات الاستيراد",
        "alerts": alerts,
        "unread_count": unread_count,
    })


def smart_import(request):
    clients = ClientBalance.objects.all().order_by("name")
    from django.db.models import Sum
    db_count = Database.objects.count()
    db_total_egp = Database.objects.aggregate(s=Sum("transfered_amount"))["s"] or 0
    cap_count = Capital.objects.count()
    cap_total_egp = Capital.objects.aggregate(s=Sum("cash_in"))["s"] or 0
    all_alerts = ImportAlert.objects.all()[:50]
    return render(request, "smart_import.html", {
        "title": "استيراد ذكي",
        "clients": clients,
        "today": dt.now().date(),
        "db_count": db_count,
        "db_total_egp": db_total_egp,
        "cap_count": cap_count,
        "cap_total_egp": cap_total_egp,
        "all_alerts": all_alerts,
    })


def clear_all_alerts(request):
    if request.method != "POST":
        return redirect("smart_import")
    user = get_current_user(request)
    if not user or not user.is_admin:
        messages.error(request, "غير مصرح")
        return redirect("smart_import")
    ImportAlert.objects.all().delete()
    messages.success(request, "تم حذف جميع التنبيهات")
    return redirect("smart_import")


@csrf_exempt
def api_smart_detect(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)
    text = request.POST.get("text", "").strip()
    content_type = request.POST.get("type", "")
    if not text:
        return JsonResponse({"error": "no text"}, status=400)

    if content_type == "balance":
        parsed = parse_balance_lines(text)
        results = []
        for item in parsed:
            results.append({
                "egp": item.get("egp", 0),
                "rate": item.get("rate", 0),
                "lyd": item.get("lyd", 0),
            })
        return JsonResponse({"results": results, "type": "balance", "count": len(results)})
    else:
        parsed = parse_whatsapp_text(text)
        results = []
        for item in parsed:
            results.append({
                "receiver_tele": item.get("receiver_tele", ""),
                "transfer_type": item.get("transfer_type", "كاش"),
                "amount_egp": item.get("amount_egp", 0),
                "exchange_rate": item.get("exchange_rate", 0),
                "amount_lyd": item.get("amount_lyd", 0),
            })
        return JsonResponse({"results": results, "type": "transfer", "count": len(results)})



@csrf_exempt
def api_extension_receive(request):
    """API endpoint for Chrome extension to send extracted messages."""
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "invalid JSON"}, status=400)

    text = data.get("text", "").strip()
    content_type = data.get("type", "auto")
    if not text:
        return JsonResponse({"error": "no text"}, status=400)

    if content_type == "auto":
        t = text.lower()
        balance_score = sum(3 if k in t else 0 for k in ["جنيه", "ج م", "القيمة", "القيمه"])
        balance_score += sum(2 if k in t else 0 for k in ["مصري", "سعر"])
        transfer_score = sum(3 if k in t else 0 for k in ["حوالة", "حواله", "فودافون", "فادفون", "فدفون", "تحويل بنكي"])
        transfer_score += sum(2 if k in t else 0 for k in ["كاش", "انستا", "صك"])
        if re.search(r'01\d{9}', text):
            transfer_score += 2
        content_type = "transfer" if transfer_score >= 2 else "balance" if balance_score >= 3 else "transfer"

    if content_type == "balance":
        parsed = parse_balance_lines(text)
        results = [{"egp": i.get("egp",0), "rate": i.get("rate",0), "lyd": i.get("lyd",0)} for i in parsed]
        return JsonResponse({"results": results, "type": "balance", "count": len(results)})
    else:
        parsed = parse_whatsapp_text(text)
        last_rate = None
        for item in parsed:
            if item.get("exchange_rate"):
                last_rate = item["exchange_rate"]
            elif last_rate:
                item["exchange_rate"] = last_rate
                if item.get("amount_egp") and last_rate > 0:
                    item["amount_lyd"] = round(item["amount_egp"] / last_rate, 2)
        results = [{
            "receiver_tele": i.get("receiver_tele",""),
            "transfer_type": i.get("transfer_type","كاش"),
            "amount_egp": i.get("amount_egp",0),
            "exchange_rate": i.get("exchange_rate",0),
            "amount_lyd": i.get("amount_lyd",0),
        } for i in parsed]
        return JsonResponse({"results": results, "type": "transfer", "count": len(results)})


@login_required_custom
def backups_list(request):
    user = get_current_user(request)
    if not user or not user.is_admin:
        messages.error(request, "ليس لديك صلاحية")
        return redirect("index")
    _ensure_backup_dir()
    files = []
    for fn in os.listdir(BACKUP_DIR):
        if fn.endswith('.csv'):
            fp = os.path.join(BACKUP_DIR, fn)
            stat = os.stat(fp)
            ftype = "حوالات" if fn.startswith("transactions") else "أرصدة"
            files.append({
                "name": fn,
                "type": ftype,
                "size": f"{stat.st_size / 1024:.1f} KB",
                "date": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    files.sort(key=lambda x: x["date"], reverse=True)
    return render(request, "backups_list.html", {
        "title": "النسخ الاحتياطية",
        "backups": files,
    })


@login_required_custom
def backup_restore(request, filename):
    user = get_current_user(request)
    if not user or not user.is_admin:
        messages.error(request, "ليس لديك صلاحية")
        return redirect("index")
    filepath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(filepath):
        messages.error(request, "الملف غير موجود")
        return redirect("backups_list")
    if filename.startswith("transactions"):
        count = _restore_transactions(filepath)
        messages.success(request, f"✅ تم استعادة {count} حوالة من النسخة الاحتياطية: {filename}")
        return redirect("transactions")
    elif filename.startswith("capital"):
        count = _restore_capital(filepath)
        messages.success(request, f"✅ تم استعادة {count} رصيد من النسخة الاحتياطية: {filename}")
        return redirect("capital_list")
    else:
        messages.error(request, "نوع ملف غير معروف")
        return redirect("backups_list")


@login_required_custom
def backup_download(request, filename):
    user = get_current_user(request)
    if not user or not user.is_admin:
        messages.error(request, "ليس لديك صلاحية")
        return redirect("index")
    filepath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(filepath):
        messages.error(request, "الملف غير موجود")
        return redirect("backups_list")
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        response = HttpResponse(f.read(), content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@login_required_custom
def backup_delete(request, filename):
    user = get_current_user(request)
    if not user or not user.is_admin:
        messages.error(request, "ليس لديك صلاحية")
        return redirect("index")
    filepath = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        messages.success(request, f"تم حذف النسخة الاحتياطية: {filename}")
    return redirect("backups_list")
